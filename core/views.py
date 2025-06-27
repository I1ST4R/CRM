from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse
from .models import Product, Order, Client, OrderItem
import json
from django.views.decorators.http import require_http_methods
from django.core.serializers import serialize
from django.db.models import F, Count, Sum
from django.contrib.admin.views.decorators import staff_member_required
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from .forms import ClientForm, ProductForm, DeliveryForm, StockMovementForm, OrderForm, OrderItemForm, OrderEditForm, DeliveryItemFormSet
from django.forms import modelformset_factory, inlineformset_factory

# Create your views here.

def hello_world(request):
    return HttpResponse("Hello World")

@require_http_methods(["GET"])
def get_product_price(request, product_id):
    try:
        product = Product.objects.get(id=product_id)
        return JsonResponse({
            'price': float(product.price),
            'stock': product.stock
        })
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)
    except Exception as e:
        print('Error:', str(e))  # Отладочная информация
        return JsonResponse({'error': str(e)}, status=500)

@require_http_methods(["GET"])
def get_customer_orders(request, customer_id):
    try:
        client = Client.objects.get(id=customer_id)
        orders = Order.objects.filter(client=client).order_by('-created_at')
        
        orders_data = []
        for order in orders:
            orders_data.append({
                'id': order.id,
                'number': order.id,  # Используем id как номер заказа
                'created_at': order.created_at.strftime('%d.%m.%Y %H:%M'),
                'status': order.get_status_display(),
                'total_amount': float(order.total_amount)
            })
            
        return JsonResponse({
            'orders': orders_data
        })
    except Client.DoesNotExist:
        return JsonResponse({'error': 'Client not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def order_report(request):
    # Получаем параметры фильтрации
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    client_id = request.GET.get('client')

    # Базовый queryset
    orders = Order.objects.all()

    # Применяем фильтры
    if date_from:
        orders = orders.filter(date__gte=date_from)
    if date_to:
        orders = orders.filter(date__lte=date_to)
    if client_id:
        orders = orders.filter(client_id=client_id)

    # Получаем статистику по статусам
    total_orders = orders.count()
    status_stats = []
    
    if total_orders > 0:
        for status, status_display in Order.STATUS_CHOICES:
            count = orders.filter(status=status).count()
            percent = round((count / total_orders) * 100, 1)
            status_stats.append((status_display, count, percent))

    # Получаем общую сумму
    total_amount = orders.aggregate(total=Sum('total_amount'))['total'] or 0

    # Получаем список всех клиентов для выпадающего списка
    clients = Client.objects.all().order_by('name')

    context = {
        'orders': orders.order_by('-date'),
        'status_stats': status_stats,
        'total_amount': total_amount,
        'clients': clients,
        'date_from': date_from,
        'date_to': date_to,
        'selected_client': client_id,
    }

    return render(request, 'order_report_user.html', context)

@login_required
def export_order_report(request):
    # Получаем параметры фильтрации
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    client_id = request.GET.get('client')

    # Базовый queryset
    orders = Order.objects.all()

    # Применяем фильтры
    if date_from:
        orders = orders.filter(date__gte=date_from)
    if date_to:
        orders = orders.filter(date__lte=date_to)
    if client_id:
        orders = orders.filter(client_id=client_id)

    # Создаем новый Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет по заказам"

    # Стили для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')

    # Заголовки
    headers = ['Номер заказа', 'Дата', 'Клиент', 'Статус', 'Сумма']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Данные
    for row, order in enumerate(orders.order_by('-date'), 2):
        ws.cell(row=row, column=1, value=f"Заказ #{order.id}")
        ws.cell(row=row, column=2, value=order.date.strftime('%d.%m.%Y'))
        ws.cell(row=row, column=3, value=order.client.name)
        ws.cell(row=row, column=4, value=order.get_status_display())
        ws.cell(row=row, column=5, value=order.total_amount)

    # Добавляем статистику
    ws.append([])  # Пустая строка
    ws.append(['Статистика по статусам'])
    ws.append(['Статус', 'Количество', 'Процент'])

    total_orders = orders.count()
    if total_orders > 0:
        for status, status_display in Order.STATUS_CHOICES:
            count = orders.filter(status=status).count()
            percent = round((count / total_orders) * 100, 1)
            ws.append([status_display, count, f"{percent}%"])

    # Общая сумма
    total_amount = orders.aggregate(total=Sum('total_amount'))['total'] or 0
    ws.append([])
    ws.append(['Общая сумма заказов:', total_amount])

    # Настройка ширины столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Создаем ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=order_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response

def clients_list(request):
    from .models import Client
    clients = Client.objects.all().order_by('-created_at')
    return render(request, 'clients_list.html', {'clients': clients})

def products_list(request):
    from .models import Product
    products = Product.objects.all().order_by('-created_at')
    return render(request, 'products_list.html', {'products': products})

def orders_list(request):
    from .models import Order
    orders = Order.objects.all().order_by('-created_at')
    return render(request, 'orders_list.html', {'orders': orders})

def deliveries_list(request):
    from .models import Delivery
    deliveries = Delivery.objects.all().order_by('-created_at')
    return render(request, 'deliveries_list.html', {'deliveries': deliveries})

def stock_list(request):
    from .models import StockMovement
    stock = StockMovement.objects.all().order_by('-date')
    return render(request, 'stock_list.html', {'stock': stock})

# --- CRUD заглушки для клиентов ---
def client_add(request):
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('clients_list')
    else:
        form = ClientForm()
    return render(request, 'client_form.html', {'form': form, 'action': 'add'})

def client_edit(request, client_id):
    client = get_object_or_404(Client, id=client_id)
    if request.method == 'POST':
        form = ClientForm(request.POST, instance=client)
        if form.is_valid():
            form.save()
            return redirect('clients_list')
    else:
        form = ClientForm(instance=client)
    return render(request, 'client_form.html', {'form': form, 'action': 'edit', 'client': client})

def client_delete(request, client_id):
    client = get_object_or_404(Client, id=client_id)
    if request.method == 'POST':
        client.delete()
        return redirect('clients_list')
    return render(request, 'client_confirm_delete.html', {'client': client})

# --- CRUD заглушки для товаров ---
def product_add(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('products_list')
    else:
        form = ProductForm()
    return render(request, 'product_form.html', {'form': form, 'action': 'add'})

def product_edit(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            return redirect('products_list')
    else:
        form = ProductForm(instance=product)
    return render(request, 'product_form.html', {'form': form, 'action': 'edit', 'product': product})

def product_delete(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        product.delete()
        return redirect('products_list')
    return render(request, 'product_confirm_delete.html', {'product': product})

# --- CRUD для заказов ---
def order_add(request):
    OrderItemFormSet = inlineformset_factory(Order, OrderItem, form=OrderItemForm, extra=1, can_delete=True)
    if request.method == 'POST':
        form = OrderForm(request.POST)
        formset = OrderItemFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            order = form.save(commit=False)
            order.date = timezone.now().date()
            order.created_by = request.user
            order.save()
            items = formset.save(commit=False)
            for item in items:
                item.order = order
                item.save()
            return redirect('orders_list')
    else:
        form = OrderForm()
        formset = OrderItemFormSet()
    return render(request, 'order_form.html', {'form': form, 'formset': formset, 'action': 'add'})

def order_edit(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    # Только изменение статуса, товары и клиент не редактируются
    if request.method == 'POST':
        form = OrderEditForm(request.POST, instance=order)
        if form.is_valid():
            form.save()
            return redirect('orders_list')
    else:
        form = OrderEditForm(instance=order)
    items = OrderItem.objects.filter(order=order)
    return render(request, 'order_edit.html', {'order': order, 'items': items, 'form': form, 'action': 'edit'})

def order_delete(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    if request.method == 'POST':
        order.delete()
        return redirect('orders_list')
    return render(request, 'order_confirm_delete.html', {'order': order})

# --- CRUD для привозов ---
def delivery_add(request):
    from .forms import DeliveryForm, DeliveryItemFormSet
    if request.method == 'POST':
        form = DeliveryForm(request.POST)
        formset = DeliveryItemFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            delivery = form.save(commit=False)
            delivery.created_by = request.user
            delivery.save()
            items = formset.save(commit=False)
            for item in items:
                item.delivery = delivery
                item.save()
            return redirect('deliveries_list')
    else:
        form = DeliveryForm()
        formset = DeliveryItemFormSet()
    return render(request, 'delivery_form.html', {'form': form, 'formset': formset, 'action': 'add'})

def delivery_edit(request, delivery_id):
    from .models import Delivery
    from .forms import DeliveryForm, DeliveryItemFormSet
    delivery = get_object_or_404(Delivery, id=delivery_id)
    if request.method == 'POST':
        form = DeliveryForm(request.POST, instance=delivery)
        formset = DeliveryItemFormSet(request.POST, instance=delivery)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            return redirect('deliveries_list')
    else:
        form = DeliveryForm(instance=delivery)
        formset = DeliveryItemFormSet(instance=delivery)
    return render(request, 'delivery_form.html', {'form': form, 'formset': formset, 'action': 'edit', 'delivery': delivery})

def delivery_delete(request, delivery_id):
    from .models import Delivery
    delivery = get_object_or_404(Delivery, id=delivery_id)
    if request.method == 'POST':
        delivery.delete()
        return redirect('deliveries_list')
    return render(request, 'delivery_confirm_delete.html', {'delivery': delivery})

# --- CRUD для движений товаров ---
def stock_add(request):
    if request.method == 'POST':
        form = StockMovementForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('stock_list')
    else:
        form = StockMovementForm()
    return render(request, 'stock_form.html', {'form': form, 'action': 'add'})

def stock_edit(request, stock_id):
    from .models import StockMovement
    stock = get_object_or_404(StockMovement, id=stock_id)
    if request.method == 'POST':
        form = StockMovementForm(request.POST, instance=stock)
        if form.is_valid():
            form.save()
            return redirect('stock_list')
    else:
        form = StockMovementForm(instance=stock)
    return render(request, 'stock_form.html', {'form': form, 'action': 'edit', 'stock': stock})

def stock_delete(request, stock_id):
    from .models import StockMovement
    stock = get_object_or_404(StockMovement, id=stock_id)
    if request.method == 'POST':
        stock.delete()
        return redirect('stock_list')
    return render(request, 'stock_confirm_delete.html', {'stock': stock})

@login_required
def main_page(request):
    user = request.user
    role = 'user'
    if user.is_superuser or user.groups.filter(name='Администраторы').exists():
        role = 'admin'
    elif user.groups.filter(name='Менеджеры').exists():
        role = 'manager'
    elif user.groups.filter(name='Кладовщики').exists():
        role = 'warehouse'
    return render(request, 'index.html', {'role': role})

def order_item_delete(request, item_id):
    item = get_object_or_404(OrderItem, id=item_id)
    order_id = item.order.id
    if request.method == 'POST':
        item.delete()
    return redirect('order_edit', order_id=order_id)
